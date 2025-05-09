# analyze_chats.py

import json
import os
import sys
from google import genai
from datetime import datetime
import time
from dotenv import load_dotenv
import pandas as pd

# --- 在脚本开头加载 .env 文件 ---
load_dotenv()
# -------------------------------

# --- 配置你的 Gemini API Key ---
API_KEY = os.environ.get('GOOGLE_API_KEY')

# --- API Key 和客户端/模型加载 ---
client = None
MODEL_NAME = 'gemini-2.5-flash-preview-04-17'

if not API_KEY:
    print("PYTHON_WARNING: GOOGLE_API_KEY not set. API calls will be skipped.", file=sys.stderr)
else:
    try:
        print(f"PYTHON_STATUS: Attempting to create Gemini client with model: {MODEL_NAME}...")
        client = genai.Client(api_key=API_KEY)
        print(f"PYTHON_STATUS: Successfully created Gemini client.")

    except Exception as e:
        client = None
        print(f"PYTHON_FATAL_ERROR: Failed to create Gemini client or initialize API.", file=sys.stderr)
        print(f"PYTHON_FATAL_ERROR_DETAIL: {e}", file=sys.stderr)
        print("PYTHON_STATUS: API calls will be skipped.")

def format_chat_transcript(messages):
    transcript = []
    for message in messages:
        transcript.append(f"[{message.get('time', '未知时间')}] {message.get('sender', '未知发送者')}: {message.get('content', '')}")
    return "\n".join(transcript)

def calculate_timing_metrics(messages):
    first_customer_time = None
    first_agent_reply_time = None
    customer_message_found = False
    first_message_time = None

    for i, message in enumerate(messages):
        if message.get('time'):
            try:
                msg_time = datetime.fromisoformat(message['time'].replace('Z', '+00:00'))
                if i == 0:
                    first_message_time = msg_time

                if message.get('sender') == 'Customer':
                    if not customer_message_found:
                        first_customer_time = msg_time
                        customer_message_found = True

                if customer_message_found and message.get('sender') == 'Agent':
                    if first_customer_time and msg_time is not None and msg_time >= first_customer_time:
                        first_agent_reply_time = msg_time
                        break

            except (ValueError, TypeError) as e:
                pass

    response_time_seconds = None
    is_qualified = "无回复"

    if first_customer_time and first_agent_reply_time:
        time_difference = first_agent_reply_time - first_customer_time
        if time_difference.total_seconds() >= 0:
            response_time_seconds = time_difference.total_seconds()
            is_qualified = "合格" if response_time_seconds <= 30 else "不合格"
        else:
            is_qualified = "异常时间"

    elif first_customer_time and not first_agent_reply_time:
        is_qualified = "无回复"

    return {
        "首次回复时长 (秒)": round(response_time_seconds, 2) if response_time_seconds is not None else None,
        "是否合格 (30秒内合格)": is_qualified,
        "初始对话时间段": first_message_time.strftime('%Y-%m-%d %H:%M:%S') if first_message_time else None
    }

def analyze_chat_with_gemini(transcript, client, model_name, chat_id):
    """调用 Gemini API 分析单个对话"""
    if client is None:
        return {
            "API分析错误": "API Key 未设置或客户端创建失败，跳过API分析。",
            "客户意图总结": "", "聊天质量点评 (基于内容)": "", "改进建议 (具体动作)": "", "潜在成交机会": "", "情绪负面评价": ""
        }

    prompt = f"""你是一个专业的聊天分析师。请分析以下客户服务对话文本，并严格按照以下 JSON 格式提取和总结信息。

请确保你的输出是一个完全有效的 JSON 对象，不要包含任何额外的文本说明、markdown 格式（如 ```json```）或其他非 JSON 内容。

JSON 对象的键（key）必须完全按照提供的中文名称。
如果某个信息无法从对话中提取或不适用，请使用空字符串 "" 或 null 表示，不要省略键。
"首次回复时长 (秒)" 和 "是否合格 (30秒内合格)" 这两项请忽略，我会在外部计算并填充。

分析项及要求（请严格按照此顺序在JSON中输出）：
1. 客户意图总结: 简洁概括客户联系客服的主要目的或问题。
2. 聊天质量点评 (基于内容): 基于对话内容，评价客服代表的沟通、问题解决、专业性等质量，提供具体例子支持评价。
3. 改进建议 (具体动作): 基于聊天质量点评，提出具体、可操作的改进措施或培训建议。
4. 潜在成交机会: 分析对话中是否存在销售、升级或其他成交的机会，并说明原因或类型。
5. 情绪负面评价: 评价客户在对话中是否表现出负面情绪（如不满、生气、沮丧等），并简要说明原因或程度。


以下是对话文本：
---对话开始---
{transcript}
---对话结束---

请输出 JSON 结果：
"""

    response = None
    analysis_result = None

    try:
        print(f"PYTHON_STATUS: Calling API for chat {chat_id} (Prompt len: {len(prompt)})...")
        start_time = time.time()
        response = client.models.generate_content(
            model=model_name,
            contents=prompt
        )
        end_time = time.time()
        print(f"PYTHON_STATUS: API call successful for chat {chat_id}, took {end_time - start_time:.2f} seconds.")

        response_text = response.text.strip()

        # --- Check prompt feedback or candidate finish reason ---
        if hasattr(response, 'prompt_feedback') and response.prompt_feedback is not None:
             prompt_feedback = response.prompt_feedback
             if hasattr(prompt_feedback, 'block_reason') and prompt_feedback.block_reason:
                 block_reason = prompt_feedback.block_reason.name
                 print(f"PYTHON_ERROR: Prompt blocked for chat {chat_id}. Reason: {block_reason}", file=sys.stderr)
                 safety_info = ""
                 if hasattr(prompt_feedback, 'safety_ratings') and prompt_feedback.safety_ratings:
                      safety_info = ", ".join([f"{s.category.name}: {s.probability.name}" for s in prompt_feedback.safety_ratings])
                      print(f"PYTHON_ERROR: Prompt safety ratings: {safety_info}", file=sys.stderr)
                 return {
                     "API分析错误": f"Prompt blocked: {block_reason}" + (f" ({safety_info})" if safety_info else ""),
                     "客户意图总结": "Prompt blocked", "聊天质量点评 (基于内容)": "", "改进建议 (具体动作)": "", "潜在成交机会": "", "情绪负面评价": ""
                 }
             if hasattr(prompt_feedback, 'safety_ratings') and prompt_feedback.safety_ratings:
                 safety_info = ", ".join([f"{s.category.name}: {s.probability.name}" for s in prompt_feedback.safety_ratings])
                 print(f"PYTHON_WARNING: Prompt safety ratings received for chat {chat_id}: {safety_info}", file=sys.stderr)

        if hasattr(response, 'candidates') and response.candidates and response.candidates[0] is not None:
             candidate = response.candidates[0]
             if hasattr(candidate, 'finish_reason') and candidate.finish_reason and candidate.finish_reason.name != 'STOP':
                  finish_reason = candidate.finish_reason.name
                  print(f"PYTHON_WARNING: Model finished with reason: {finish_reason} for chat {chat_id}.", file=sys.stderr)
                  reason_detail = ""
                  if hasattr(candidate, 'safety_ratings') and candidate.safety_ratings:
                       reason_detail = " Safety Ratings: " + ", ".join([f"{s.category.name}: {s.probability.name}" for s in candidate.safety_ratings])
                       print(f"PYTHON_WARNING: Candidate safety ratings: {reason_detail}", file=sys.stderr)

                  if not response_text:
                      return {
                         "API分析错误": f"Model finished with reason: {finish_reason}" + (f" ({reason_detail})" if reason_detail else "") + ". No text generated.",
                         "客户意图总结": "Generation failed", "聊天质量点评 (基于内容)": "", "改进建议 (具体动作)": "", "潜在成交机会": "", "情绪负面评价": ""
                     }

        # --- JSON Parsing Attempt ---
        if not response_text:
            print(f"PYTHON_ERROR: API call successful but returned empty text for chat {chat_id}.", file=sys.stderr)
            return {
                "API分析错误": "API returned empty text.",
                "客户意图总结": "Empty response", "聊天质量点评 (基于内容)": "", "改进建议 (具体动作)": "", "潜在成交机会": "", "情绪负面评价": ""
            }

        try:
            # First attempt: direct parse
            analysis_result = json.loads(response_text)
        except json.JSONDecodeError as json_e:
            # If direct parse fails, print error and snippet
            print(f"PYTHON_ERROR: JSON Decode Error for chat {chat_id}: {json_e}", file=sys.stderr)
            print(f"PYTHON_ERROR_SNIPPET: Attempted to parse: {response_text[:500]}...", file=sys.stderr)
            # Try robust parsing as a fallback
            try:
                clean_text = response_text.replace('```json', '').replace('```', '').strip()
                analysis_result = json.loads(clean_text)
                print(f"PYTHON_WARNING: Successfully parsed after cleanup for chat {chat_id}.", file=sys.stderr)
            except json.JSONDecodeError:
                 print(f"PYTHON_FATAL_ERROR: Robust JSON parsing failed for chat {chat_id}.", file=sys.stderr)
                 raise # Re-raise if robust parsing fails


        # --- Validate parsed result ---
        if not isinstance(analysis_result, dict):
             raise ValueError("Parsed JSON result is not a valid dictionary.")

        expected_keys = ["客户意图总结", "聊天质量点评 (基于内容)", "改进建议 (具体动作)", "潜在成交机会", "情绪负面评价"]
        for key in expected_keys:
            if analysis_result.get(key) is None:
                 analysis_result[key] = ""

        # --- NEW: Print parsed result here if successful ---
        # This is the data for the *current* chat displayed in the frontend
        try:
             print(f"PYTHON_PARSED_CHAT_RESULT: {json.dumps(analysis_result, ensure_ascii=False)}")
        except Exception as e:
             print(f"PYTHON_ERROR: Failed to serialize parsed result for stream for chat {chat_id}: {e}", file=sys.stderr)
             # Still return the parsed result even if streaming fails
        # ---------------------------------------------------

        return analysis_result # Return the successfully parsed result

    except Exception as e:
        print(f"PYTHON_ERROR: Exception caught during analysis for chat {chat_id}: {type(e).__name__} - {e}", file=sys.stderr)
        # Try to include any available safety/finish info
        safety_info_parts = []
        if 'response' in locals() and response is not None:
             if hasattr(response, 'prompt_feedback') and response.prompt_feedback is not None:
                 prompt_feedback = response.prompt_feedback
                 if hasattr(prompt_feedback, 'safety_ratings') and prompt_feedback.safety_ratings:
                      safety_info_parts.append("Prompt Safety: " + ", ".join([f"{s.category.name}: {s.probability.name}" for s in prompt_feedback.safety_ratings]))
                 if hasattr(prompt_feedback, 'block_reason') and prompt_feedback.block_reason:
                       safety_info_parts.append(f"Prompt Blocked: {prompt_feedback.block_reason.name}")

             if hasattr(response, 'candidates') and response.candidates and response.candidates[0] is not None:
                  candidate = response.candidates[0]
                  if hasattr(candidate, 'finish_reason') and candidate.finish_reason and candidate.finish_reason.name != 'STOP':
                       finish_reason = candidate.finish_reason.name
                       safety_info_parts.append(f"Finish Reason: {finish_reason}")
                       if hasattr(candidate, 'safety_ratings') and candidate.safety_ratings:
                           safety_info_parts.append("Candidate Safety: " + ", ".join([f"{s.category.name}: {s.probability.name}" for s in candidate.safety_ratings]))

        safety_info = "; ".join(safety_info_parts) if safety_info_parts else ""

        error_message_detail = str(e)
        if safety_info:
             error_message_detail = f"{error_message_detail} ({safety_info})"

        # Print traceback for unexpected errors (not handled JSONDecodeErrors)
        if not (isinstance(e, json.JSONDecodeError) and 'Successfully parsed after cleanup' in str(e)):
             import traceback
             traceback.print_exc(file=sys.stderr)


        return {
            "API分析错误": f"{type(e).__name__}: {error_message_detail}",
            "客户意图总结": "Analysis error", "聊天质量点评 (基于内容)": "", "改进建议 (具体动作)": "", "潜在成交机会": "", "情绪负面评价": ""
        }


# --- 将核心分析和保存逻辑封装到函数中 ---
def run_analysis_process(cleaned_input_file, final_output_file_excel, client_instance, model_name_str, limit=None, print_results_to_console=True):
    print(f"PYTHON_STATUS: Loading cleaned data from {cleaned_input_file}...")
    try:
        with open(cleaned_input_file, 'r', encoding='utf-8') as f:
            cleaned_chats = json.load(f)
        print(f"PYTHON_STATUS: Successfully loaded cleaned data. Found {len(cleaned_chats)} chats.")
    except FileNotFoundError:
        print(f"PYTHON_FATAL_ERROR: Cleaned data file not found: {cleaned_input_file}.", file=sys.stderr)
        return []
    except json.JSONDecodeError:
        print(f"PYTHON_FATAL_ERROR: Failed to parse cleaned data file {cleaned_input_file}. Check JSON format.", file=sys.stderr)
        return []
    except Exception as e:
        print(f"PYTHON_FATAL_ERROR: Unknown error loading cleaned data: {e}", file=sys.stderr)
        return []

    analyzed_results = []
    chats_to_process = cleaned_chats[:limit] if limit is not None else cleaned_chats
    total_chats_to_process = len(chats_to_process)
    print(f"PYTHON_STATUS: Starting analysis for {total_chats_to_process} chats.")

    successful_api_calls = 0
    skipped_api_calls = 0

    print("PYTHON_STATUS: Beginning chat analysis loop...")
    for i, chat in enumerate(chats_to_process):
        chat_id = chat.get('chat_id', f'UnknownID_{i+1}')

        # 打印总体进度 (更精确地报告当前处理的是第几条总共多少条)
        print(f"PYTHON_STATUS: Processing chat [{i + 1}/{total_chats_to_process}] (ID: {chat_id})")

        messages = chat.get('messages', [])

        if not messages:
            print(f"PYTHON_STATUS: Chat {chat_id} has no valid messages, skipping analysis.")
            analyzed_results.append({
                 "chat_id": chat_id,
                 "客户意图总结": "", "聊天质量点评 (基于内容)": "无有效消息", "改进建议 (具体动作)": "", "首次回复时长 (秒)": None, "是否合格 (30秒内合格)": "无回复", "潜在成交机会": "", "情绪负面评价": ""
            })
            skipped_api_calls += 1
            continue

        timing_metrics = calculate_timing_metrics(messages)
        transcript = format_chat_transcript(messages)

        # 调用 API 分析函数
        analysis_content = analyze_chat_with_gemini(transcript, client_instance, model_name_str, chat_id)
        # analyze_chat_with_gemini function now prints the parsed JSON with PYTHON_PARSED_CHAT_RESULT prefix

        # Combine results for the final Excel/JSON list
        chat_analysis = {
            "chat_id": chat_id, # --- Ensure chat_id is included here ---
            "客户姓名": chat.get('customer', {}).get('name', '') if chat.get('customer', {}).get('name', '') else '[无信息]',
            "初始对话时间段": timing_metrics.get("初始对话时间段"),
            "客户意图总结": analysis_content.get("客户意图总结", "") if analysis_content.get("客户意图总结", "") else "[无信息]",
            "聊天质量点评 (基于内容)": analysis_content.get("聊天质量点评 (基于内容)", "") if analysis_content.get("聊天质量点评 (基于内容)", "") else "[无信息]",
            "改进建议 (具体动作)": analysis_content.get("改进建议 (具体动作)", "") if analysis_content.get("改进建议 (具体动作)", "") else "[无信息]",
            "潜在成交机会": analysis_content.get("潜在成交机会", "") if analysis_content.get("潜在成交机会", "") else "[无信息]",
            "情绪负面评价": analysis_content.get("情绪负面评价", "") if analysis_content.get("情绪负面评价", "") else "[无信息]",
            "首次回复时长 (秒)": timing_metrics["首次回复时长 (秒)"],
            "是否合格 (30秒内合格)": timing_metrics["是否合格 (30秒内合格)"]
        }

        # Include API error info if present
        if "API分析错误" in analysis_content:
             chat_analysis["API分析错误"] = analysis_content["API分析错误"]
             skipped_api_calls += 1
             print(f"PYTHON_STATUS: Analysis failed for chat {chat_id} with error: {analysis_content['API分析错误']}.", file=sys.stderr) # Log error status
        else:
            successful_api_calls += 1
            print(f"PYTHON_STATUS: Analysis result processed for chat {chat_id}.") # Log success status


        analyzed_results.append(chat_analysis)

        if client_instance and "API分析错误" not in analysis_content:
             time.sleep(6.0) # Delay to respect rate limits

    print(f"\nPYTHON_STATUS: Finished analysis loop for {total_chats_to_process} chats.")
    print(f"PYTHON_STATUS: Summary: Successful API calls: {successful_api_calls}, Skipped/Failed: {skipped_api_calls}, Total Processed: {total_chats_to_process}")

    # --- NEW: Calculate Overall Summary Metrics ---
    print("\nPYTHON_STATUS: Calculating overall analysis metrics...")
    total_chats_analyzed = len(analyzed_results)
    summary_data = {}
    if total_chats_analyzed > 0:
        # Basic Quantitative Summary
        qualified_count = sum(1 for r in analyzed_results if r.get("是否合格 (30秒内合格)") == "合格")
        no_reply_count = sum(1 for r in analyzed_results if r.get("是否合格 (30秒内合格)") == "无回复")
        api_error_count = sum(1 for r in analyzed_results if r.get("API分析错误") != "" and r.get("API分析错误") is not None)

        qualified_percentage = (qualified_count / total_chats_analyzed) * 100 if total_chats_analyzed > 0 else 0

        # Calculate average response time for *qualified* chats (those with a response time)
        # Filter out None values before calculating average
        response_times = [r.get("首次回复时长 (秒)") for r in analyzed_results if r.get("首次回复时长 (秒)") is not None]
        average_response_time = sum(response_times) / len(response_times) if response_times else None

        summary_data = {
            "total_chats_processed": total_chats_to_process,
            "total_chats_with_results": total_chats_analyzed,
            "successful_api_calls": successful_api_calls,
            "skipped_api_calls": skipped_api_calls,
            "qualified_response_count": qualified_count,
            "qualified_response_percentage": round(qualified_percentage, 2),
            "no_reply_count": no_reply_count,
            "api_error_count": api_error_count,
            "average_response_time_seconds": round(average_response_time, 2) if average_response_time is not None else None,
        }

        # --- Print machine-readable summary for Node.js to capture (optional but kept for consistency) ---
        # Node.js can choose to capture this or the final AI summary.
        # print(f"PYTHON_SUMMARY_JSON:{json.dumps(summary_data, ensure_ascii=False)}")

        # --- Also print a human-readable summary to stderr for visibility (optional but helpful) ---
        print("\n--- Chat Analysis Overall Summary (Metrics) ---", file=sys.stderr)
        print(f"Processed Chats: {total_chats_to_process}", file=sys.stderr)
        print(f"Chats with Results: {total_chats_analyzed}", file=sys.stderr)
        print(f"Successful API Calls: {successful_api_calls}", file=sys.stderr)
        print(f"Skipped/Failed API Calls: {skipped_api_calls}", file=sys.stderr)
        print(f"Chats Replied within 30s: {qualified_count} ({qualified_percentage:.2f}%)", file=sys.stderr)
        print(f"Chats with No Agent Reply: {no_reply_count}", file=sys.stderr)
        print(f"Chats with API Errors: {api_error_count}", file=sys.stderr)
        if average_response_time is not None:
            print(f"Average Response Time (for replied chats): {average_response_time:.2f} seconds", file=sys.stderr)
        else:
             print("Average Response Time: N/A (No chats with agent replies)", file=sys.stderr)
        print("-------------------------------------", file=sys.stderr)

    else:
        print("PYTHON_STATUS: No analysis results to summarize metrics from.", file=sys.stderr)

    # --- NEW: Generate Overall Summary using AI ---
    print("\nPYTHON_STATUS: Requesting overall analysis summary from AI...")
    overall_summary_text = "整体总结生成跳过：API 客户端未初始化或无分析结果。"

    # Only attempt AI summary if client is available and there are results to analyze
    if client_instance is not None and total_chats_analyzed > 0:
        try:
            # Prepare the summary data for the prompt
            summary_data_formatted = "关键汇总指标：\n"
            for key, value in summary_data.items():
                 # Format value appropriately
                 display_value = f"{value:.2f}%" if isinstance(value, float) and key == "qualified_response_percentage" else value
                 summary_data_formatted += f"- {key}：{display_value}\n"

            # Construct the prompt for the overall summary
            # Reference the types of qualitative analysis performed earlier
            overall_summary_prompt = f"""
作为一名专业的聊天分析师，基于以下关键汇总指标和对每条对话内容的分析，请提供本次对话批次的整体表现总结：

{summary_data_formatted}

请评估整体聊天表现，直击要点，并指出存在的问题。总结应涵盖：
    -   客户响应效率（合格率、平均回复时长）
    -   客服沟通质量的整体评估（亮点与不足）
    -   主要客户需求和潜在业务机会
    -   需关注的风险点（如负面情绪对话比例、AI 分析错误）
    -   关键改进方向建议
    -   总结必须简短精炼，突出核心发现。

请直接输出总结文本：
"""

            print(f"PYTHON_STATUS: Calling AI for overall summary (Prompt len: {len(overall_summary_prompt)})...")
            start_time = time.time()
            overall_response = client_instance.models.generate_content(
                model=model_name_str,
                contents=overall_summary_prompt
            )
            end_time = time.time()
            print(f"PYTHON_STATUS: Overall summary AI call successful, took {end_time - start_time:.2f} seconds.")

            overall_summary_text = overall_response.text.strip()

            # Check for prompt feedback or finish reason for the summary call
            if hasattr(overall_response, 'prompt_feedback') and overall_response.prompt_feedback is not None:
                 prompt_feedback = overall_response.prompt_feedback
                 if hasattr(prompt_feedback, 'block_reason') and prompt_feedback.block_reason:
                      block_reason = prompt_feedback.block_reason.name
                      overall_summary_text = f"整体总结生成失败：AI Prompt Blocked. Reason: {block_reason}"
                      safety_info = ""
                      if hasattr(prompt_feedback, 'safety_ratings') and prompt_feedback.safety_ratings:
                           safety_info = ", ".join([f"{s.category.name}: {s.probability.name}" for s in prompt_feedback.safety_ratings])
                           overall_summary_text += (f" ({safety_info})" if safety_info else "")
                      print(overall_summary_text, file=sys.stderr) # Log error
                 elif hasattr(prompt_feedback, 'safety_ratings') and prompt_feedback.safety_ratings:
                      safety_info = ", ".join([f"{s.category.name}: {s.probability.name}" for s in prompt_feedback.safety_ratings])
                      print(f"PYTHON_WARNING: Overall summary prompt safety ratings: {safety_info}", file=sys.stderr)

            if hasattr(overall_response, 'candidates') and overall_response.candidates and overall_response.candidates[0] is not None:
                 candidate = overall_response.candidates[0]
                 if hasattr(candidate, 'finish_reason') and candidate.finish_reason and candidate.finish_reason.name != 'STOP':
                      finish_reason = candidate.finish_reason.name
                      reason_detail = " Safety Ratings: " + ", ".join([f"{s.category.name}: {s.probability.name}" for s in candidate.safety_ratings]) if hasattr(candidate, 'safety_ratings') and candidate.safety_ratings else ""
                      overall_summary_text = f"整体总结生成警告：AI 完成原因非STOP ({finish_reason}{reason_detail}). 生成内容可能不完整或异常。\n" + overall_summary_text # Prepend warning
                      print(overall_summary_text, file=sys.stderr) # Log warning

            # --- Print the AI-generated overall summary (still useful for Node.js stream) ---
            print("\n--- AI Generated Overall Analysis Summary (Console/Stderr) ---", file=sys.stderr) # Changed prefix to stderr
            print(f"PYTHON_OVERALL_SUMMARY:{overall_summary_text}", file=sys.stderr) # Use specific prefix for stderr
            print("---------------------------------------------", file=sys.stderr)

        except Exception as e:
            overall_summary_text = f"整体总结生成失败：发生异常 - {type(e).__name__} - {e}"
            print(overall_summary_text, file=sys.stderr) # Log error
            import traceback
            traceback.print_exc(file=sys.stderr)

    else:
        # Print message if AI summary was skipped
        print(f"PYTHON_OVERALL_SUMMARY:{overall_summary_text}")


    # --- 保存分析结果到 Excel 文件 ---
    print(f"\nPYTHON_STATUS: Attempting to save analysis results to Excel file: {final_output_file_excel}...")
    try:
        df = pd.DataFrame(analyzed_results)
        column_order = [
            "chat_id",
            "客户姓名",
            "初始对话时间段",
            "客户意图总结",
            "聊天质量点评 (基于内容)",
            "改进建议 (具体动作)",
            "首次回复时长 (秒)",
            "是否合格 (30秒内合格)",
            "潜在成交机会",
            "情绪负面评价",
            "API分析错误"
        ]
        df = df.reindex(columns=column_order)
        df = df.fillna("")

        # Use ExcelWriter to write multiple sheets
        with pd.ExcelWriter(final_output_file_excel, engine='openpyxl') as writer:
            # Write the detailed analysis results to the first sheet
            df.to_excel(writer, sheet_name='Detailed Analysis', index=False)

            # 优化列宽和行高（自动换行，宽度适中）
            worksheet = writer.sheets['Detailed Analysis']
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # 获取列字母
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                # 设置较小的基础宽度，适合自动换行
                adjusted_width = min((max_length + 2) * 0.9, 30)  # 限制最大宽度为30
                worksheet.column_dimensions[column].width = adjusted_width
            worksheet.row_dimensions[1].height = 28  # 表头行高

            # 设置自动换行和内容居中
            from openpyxl.styles import Alignment
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            # --- NEW: Write the AI-generated overall summary to a second sheet ---
            summary_lines = overall_summary_text.split('\n')
            summary_df = pd.DataFrame(summary_lines, columns=['Overall Analysis Summary'])
            summary_df.to_excel(writer, sheet_name='Overall Summary', index=False)

            # 美化 Overall Summary 工作表
            summary_ws = writer.sheets['Overall Summary']
            from openpyxl.styles import Alignment, Font, PatternFill
            # 设置表头加粗、底色
            header_cell = summary_ws['A1']
            header_cell.font = Font(bold=True, size=13)
            header_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            # 设置内容区域字体、行高、适当列宽
            for row in summary_ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(size=12)
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            # 自动调整列宽（尽量大，适合长文本）
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in summary_ws['A'])
            summary_ws.column_dimensions['A'].width = min((max_length + 8) * 1.2, 120)  # 最大宽度120
            # 设置每行高度更大，适合长文本自动换行
            for row in range(2, summary_ws.max_row + 1):
                summary_ws.row_dimensions[row].height = 38

        print(f"PYTHON_STATUS: Analysis results and overall summary successfully saved to Excel file: {final_output_file_excel}")
    except ImportError:
        print(f"PYTHON_FATAL_ERROR: Saving to Excel requires pandas and openpyxl. Install with 'pip install pandas openpyxl'.", file=sys.stderr)
    except IOError as e:
        print(f"PYTHON_FATAL_ERROR: IO error writing file {final_output_file_excel}: {e}", file=sys.stderr)
        print("PYTHON_FATAL_ERROR: Please ensure the file is not open and you have write permissions.", file=sys.stderr)
    except Exception as e:
        print(f"PYTHON_FATAL_ERROR: Unknown error saving analysis results to Excel: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)

    return analyzed_results # Optionally return summary_data and overall_summary_text as well if needed by caller

# --- 原来的 __main__ 块修改为独立运行时的逻辑 ---
if __name__ == '__main__':
    print("--- analyze_chats.py running as independent script ---")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_filename = input(f"Enter the cleaned chats JSON filename (in script directory, e.g., cleaned_chats.json): ")
    output_filename = input(f"Enter the output Excel filename (in script directory, e.g., analyzed_chats_results.xlsx): ")
    limit_input_str = input(f"Enter number of chats to process (leave empty or non-number for all): ")

    limit_value = None
    try:
        if limit_input_str.strip():
            limit_value = int(limit_input_str)
            if limit_value < 0: limit_value = None
    except ValueError: pass

    input_file = os.path.join(script_dir, input_filename)
    output_file = os.path.join(script_dir, output_filename)

    # When running independently, we might want pretty JSON print
    # Setting print_results_to_console=True will trigger the JSON dump with prefix
    # If you want pretty print in the terminal, you might need conditional print logic in run_analysis_process
    run_analysis_process(input_file, output_file, client, MODEL_NAME, limit=limit_value, print_results_to_console=True)

    print("\n--- analyze_chats.py independent run finished ---")